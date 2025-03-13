from abc import ABC, abstractmethod
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
import calendar
from sqlalchemy import select
from components import session
from components.dbmodel import Duty
from components.rules.uatshift import ShiftCalendar


class Report(ABC):
    @abstractmethod
    def display(self):
        pass

    @abstractmethod
    def export(self):
        pass


class DutyReport(Report):
    def __init__(self, year, month):
        self.year = year
        self.month = month

    def display(self):
        _, days_in_month = calendar.monthrange(self.year, self.month)
        print(f"{self.year}年{self.month}月值班表：")
        rows = self.generate_report_rows()
        for d in range(1, days_in_month + 1):
            print(f"|{self.year}-{self.month}-{d}".ljust(10), end="|")
            for row in rows:
                if row.date != date(self.year, self.month, d):
                    continue
                if row.type == "OperationRuleMiddle":
                    print(f"{row.employee}(中,夜)".ljust(12), end="|")
                elif row.type == "OperationRuleNight":
                    ignore = False
                    for item in rows:
                        if (item.type == "OperationRuleMiddle"
                                and item.employee == row.employee
                                and row.date == item.date):
                            ignore = True
                            break
                    if not ignore:
                        print(f"{row.employee}(夜)".ljust(12), end="|")
                else:
                    print(f"{row.employee}".ljust(12), end="|")
            print()


    def export(self):
        print("DutyReport export")

    def generate_report_rows(self):
        _, days_in_month = calendar.monthrange(self.year, self.month)
        start = date(self.year, self.month, 1)
        end = date(self.year, self.month, days_in_month)
        stmt = select(Duty).where(Duty.date.between(start, end))
        rows = session.execute(stmt).scalars().all()

        return rows if rows else list()


class RenderExcel:
    def __init__(self, excel_file):
        self.excel_file = excel_file

    def render(self):
        # 加载 Excel 文件
        wb = load_workbook(self.excel_file)
        ws = wb.active

        # 设定每列列宽
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 9
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['F'].width = 20

        # 定义绿色底纹样式
        green_fill = PatternFill(start_color="244062", end_color="244062", fill_type="solid")

        # 设置某一行（例如第 2 行）的底纹为绿色
        font_white = Font(color="FFFFFF")
        row_to_highlight = 1  # Excel 行号从 1 开始
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row_to_highlight, column=col).fill = green_fill
            ws.cell(row=row_to_highlight, column=col).font = font_white

        # 定义边框样式
        thin_border = Border(
            left=Side(style='thin'),  # 左边框
            right=Side(style='thin'),  # 右边框
            top=Side(style='thin'),  # 上边框
            bottom=Side(style='thin')  # 下边框
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        # 为整个数据区域添加边框
        blue_fill = PatternFill(start_color="C5D9F1", end_color="C5D9F1", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.fill = blue_fill
                cell.alignment = center_alignment

        # 标注双休日及投产日
        when_start = datetime.strptime(ws['A'][1].value, "%Y年%m月%d日")
        when_end = datetime.strptime(ws['A'][ws.max_row - 1].value, "%Y年%m月%d日")
        shift_calendar = ShiftCalendar(when_start, when_end)
        inproduct = shift_calendar.inproduct_days
        holidays = shift_calendar.holidays
        inproduct_fill = PatternFill(start_color="FABF8F", end_color="FABF8F", fill_type="solid")
        holiday_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        for cell in ws['A']:
            if cell.row == 1:
                continue
            duty_date = datetime.strptime(cell.value, "%Y年%m月%d日").date()
            if duty_date in holidays:
                for col in range(1, ws.max_column + 1):
                    ws.cell(cell.row, column=col).fill = holiday_fill
            if duty_date in inproduct:
                for col in range(1, ws.max_column + 1):
                    ws.cell(cell.row, column=col).fill = inproduct_fill

        wb.save(self.excel_file)
